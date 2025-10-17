#!/usr/bin/env python3
"""
Create BOQ Template Excel File
Generates a properly formatted Excel template with formulas and validation
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def create_boq_template():
    """Create a formatted BOQ template Excel file"""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Template"
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # SECTION 1: PROJECT INFORMATION HEADER (Rows 1-15)
    project_info = [
        ("Project Name", "Residential House Construction"),
        ("Project Type", "Residential"),
        ("Location", "Manila, Philippines"),
        ("Client", "John Doe"),
        ("Contractor", "PowerMason Construction"),
        ("Lot Size (sqm)", 150),
        ("Floor Area (sqm)", 120),
        ("Project Category", "Private"),
        ("Complexity Level", "Mid Range"),
        ("Role/Type", "General Contractor"),
        ("Date Prepared", "2024-01-15"),
        ("Prepared By", "Project Manager"),
        ("Project Value (PHP)", "=SUM(G18:G1000)"),
        ("Cost per sqm (PHP)", "=B13/B6"),
        ("Total Items", "=COUNTA(A18:A1000)")
    ]
    
    # Write project information
    for i, (label, value) in enumerate(project_info, 1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        
        # Style the label
        ws[f'A{i}'].font = subheader_font
        ws[f'A{i}'].fill = subheader_fill
        ws[f'A{i}'].border = border
        
        # Style the value
        ws[f'B{i}'].border = border
        if isinstance(value, str) and value.startswith('='):
            ws[f'B{i}'].font = Font(italic=True)
    
    # Add empty row
    ws[f'A16'] = ""
    
    # SECTION 2: BOQ ITEMS TABLE HEADER (Row 17)
    headers = [
        "Item #", "Description", "Section/Category", "UOM", "Quantity", 
        "Unit Cost", "Total Cost", "Material Cost", "Labor Cost", 
        "Equipment Cost", "Subcontractor Cost", "Dependencies", "Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Sample data rows (18-25)
    sample_data = [
        [1, "Excavation for Foundation", "Foundation", "cu.m", 15, 500, "=E18*F18", 300, 150, 50, 0, "", "Standard excavation depth 1.5m"],
        [2, "Concrete Foundation", "Foundation", "cu.m", 12, 8000, "=E19*F19", 5000, 2500, 500, 0, "1", "Ready-mix concrete"],
        [3, "Structural Steel Frame", "Structure", "kg", 2000, 120, "=E20*F20", 8000, 3000, 1000, 0, "2", "Grade 40 steel"],
        [4, "Concrete Slab", "Structure", "sqm", 120, 1500, "=E21*F21", 800, 500, 200, 0, "3", "6-inch thick slab"],
        [5, "Electrical Installation", "MEP", "lot", 1, 25000, "=E22*F22", 15000, 8000, 2000, 0, "4", "Complete electrical system"],
        [6, "Plumbing Installation", "MEP", "lot", 1, 18000, "=E23*F23", 12000, 5000, 1000, 0, "4", "Complete plumbing system"],
        [7, "Floor Tiles", "Finishing", "sqm", 120, 800, "=E24*F24", 600, 150, 50, 0, "4", "Ceramic tiles"],
        [8, "Painting Works", "Finishing", "sqm", 300, 200, "=E25*F25", 100, 80, 20, 0, "7", "Interior and exterior"]
    ]
    
    for row_idx, row_data in enumerate(sample_data, 18):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if isinstance(value, str) and value.startswith('='):
                cell.font = Font(italic=True)
    
    # TOTAL ROW (Row 26)
    total_row = [
        "TOTAL", "", "", "", "=SUM(E18:E1000)", "", "=SUM(G18:G1000)", 
        "=SUM(H18:H1000)", "=SUM(I18:I1000)", "=SUM(J18:J1000)", "=SUM(K18:K1000)", "", ""
    ]
    
    for col, value in enumerate(total_row, 1):
        cell = ws.cell(row=26, column=col, value=value)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        cell.border = border
        if isinstance(value, str) and value.startswith('='):
            cell.font = Font(bold=True, italic=True)
    
    # Add data validation for UOM column (D)
    uom_validation = DataValidation(
        type="list",
        formula1='"sqm,cu.m,pcs,kg,lot,lm,m,set,unit"',
        allow_blank=True
    )
    ws.add_data_validation(uom_validation)
    uom_validation.add(f"D18:D1000")
    
    # Add data validation for Project Category (B8)
    category_validation = DataValidation(
        type="list",
        formula1='"Public,Private,Renovation,New Build"',
        allow_blank=True
    )
    ws.add_data_validation(category_validation)
    category_validation.add("B8")
    
    # Add data validation for Complexity Level (B9)
    complexity_validation = DataValidation(
        type="list",
        formula1='"Low End,Mid Range,High End"',
        allow_blank=True
    )
    ws.add_data_validation(complexity_validation)
    complexity_validation.add("B9")
    
    # Add data validation for Role/Type (B10)
    role_validation = DataValidation(
        type="list",
        formula1='"General Contractor,Subcontractor"',
        allow_blank=True
    )
    ws.add_data_validation(role_validation)
    role_validation.add("B10")
    
    # Set column widths
    column_widths = [8, 30, 15, 8, 10, 12, 12, 12, 12, 12, 15, 15, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze panes at row 17
    ws.freeze_panes = "A17"
    
    # Add instructions as text in empty cells (simpler approach)
    ws["C1"] = "← Enter project name here"
    ws["C10"] = "← Select GC or Subcontractor"
    ws["N18"] = "← Enter dependencies (e.g., '1,3')"
    
    # Save the file
    wb.save("static/templates/BOQ_Template.xlsx")
    print("BOQ Template Excel file created successfully!")

if __name__ == "__main__":
    create_boq_template()
