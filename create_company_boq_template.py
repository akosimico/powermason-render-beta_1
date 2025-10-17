import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_company_boq_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Company BOQ Template"

    # --- Section 1: Project Information Header (Rows 1-5) ---
    # Row 1: Project
    ws["B1"] = "Project"
    ws["C1"] = ": [Enter project name here]"
    ws["B1"].font = Font(bold=True)
    
    # Row 2: Location
    ws["B2"] = "Location"
    ws["C2"] = ": [Enter project location here]"
    ws["B2"].font = Font(bold=True)
    
    # Row 3: Package
    ws["B3"] = "Package"
    ws["C3"] = ": [Enter work type/project scope here]"
    ws["B3"].font = Font(bold=True)
    
    # Row 3: Contractor info
    ws["J3"] = "Name of Contractor: [Enter contractor name]"
    ws["J3"].font = Font(bold=True)
    
    # Row 4: Proposal No
    ws["J4"] = "Proposal No: [Enter proposal number]"
    ws["J4"].font = Font(bold=True)
    
    # Row 5: BOQ Header
    ws["B5"] = "BILL OF QUANTITIES: DETAILED BREAKDOWN"
    ws["B5"].font = Font(bold=True, size=14)
    
    # Row 6-7: Empty rows for spacing
    ws["B6"] = ""
    ws["B7"] = ""

    # --- Section 2: BOQ Table Header (Row 8-9) ---
    # Headers
    ws["B8"] = "ITEM/DESCRIPTION"
    ws["G8"] = "UNIT"
    ws["H8"] = "QUANTITY"
    ws["I8"] = "UNIT COST (MATERIAL)"
    ws["J8"] = "UNIT COST (LABOR)"
    ws["L8"] = "AMOUNT"
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in ['B', 'G', 'H', 'I', 'J', 'L']:
        cell = ws[f"{col}8"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Row 9: Empty or additional headers
    ws["B9"] = ""

    # --- Section 3: Sample BOQ Items (Starting Row 10) ---
    # Section Header
    ws["B10"] = "GENERAL REQUIREMENTS"
    ws["B10"].font = Font(bold=True, size=12)
    ws["B10"].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Sample items
    sample_items = [
        {
            "description": "Site Clearing and Preparation",
            "unit": "sq.m",
            "quantity": 100,
            "material_cost": 120,
            "labor_cost": 80
        },
        {
            "description": "Excavation for Foundation",
            "unit": "cu.m",
            "quantity": 50,
            "material_cost": 500,
            "labor_cost": 300
        },
        {
            "description": "Concrete Footing",
            "unit": "cu.m",
            "quantity": 20,
            "material_cost": 4000,
            "labor_cost": 1200
        },
        {
            "description": "Reinforcement Steel",
            "unit": "kg",
            "quantity": 500,
            "material_cost": 80,
            "labor_cost": 25
        }
    ]
    
    # Add sample items
    for i, item in enumerate(sample_items, start=11):
        row = i
        ws[f"B{row}"] = item["description"]
        ws[f"G{row}"] = item["unit"]
        ws[f"H{row}"] = item["quantity"]
        ws[f"I{row}"] = item["material_cost"]
        ws[f"J{row}"] = item["labor_cost"]
        ws[f"L{row}"] = f"=(I{row}+J{row})*H{row}"  # Formula: (Material + Labor) × Quantity
        
        # Format numbers
        ws[f"H{row}"].number_format = '#,##0'
        ws[f"I{row}"].number_format = '#,##0.00'
        ws[f"J{row}"].number_format = '#,##0.00'
        ws[f"L{row}"].number_format = '#,##0.00'
    
    # Subtotal row
    subtotal_row = 11 + len(sample_items)
    ws[f"B{subtotal_row}"] = "Subtotal (General Requirements)"
    ws[f"B{subtotal_row}"].font = Font(bold=True)
    ws[f"L{subtotal_row}"] = f"=SUM(L11:L{subtotal_row-1})"
    ws[f"L{subtotal_row}"].number_format = '#,##0.00'
    ws[f"L{subtotal_row}"].font = Font(bold=True)
    
    # Add more sections
    # SITE WORKS section
    site_start = subtotal_row + 2
    ws[f"B{site_start}"] = "SITE WORKS"
    ws[f"B{site_start}"].font = Font(bold=True, size=12)
    ws[f"B{site_start}"].fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    site_items = [
        {
            "description": "Concrete Slab on Grade",
            "unit": "sq.m",
            "quantity": 200,
            "material_cost": 800,
            "labor_cost": 400
        },
        {
            "description": "Drainage System",
            "unit": "lm",
            "quantity": 150,
            "material_cost": 200,
            "labor_cost": 150
        }
    ]
    
    for i, item in enumerate(site_items, start=site_start + 1):
        row = i
        ws[f"B{row}"] = item["description"]
        ws[f"G{row}"] = item["unit"]
        ws[f"H{row}"] = item["quantity"]
        ws[f"I{row}"] = item["material_cost"]
        ws[f"J{row}"] = item["labor_cost"]
        ws[f"L{row}"] = f"=(I{row}+J{row})*H{row}"
        
        # Format numbers
        ws[f"H{row}"].number_format = '#,##0'
        ws[f"I{row}"].number_format = '#,##0.00'
        ws[f"J{row}"].number_format = '#,##0.00'
        ws[f"L{row}"].number_format = '#,##0.00'
    
    # Site works subtotal
    site_subtotal_row = site_start + 1 + len(site_items)
    ws[f"B{site_subtotal_row}"] = "Subtotal (Site Works)"
    ws[f"B{site_subtotal_row}"].font = Font(bold=True)
    ws[f"L{site_subtotal_row}"] = f"=SUM(L{site_start+1}:L{site_subtotal_row-1})"
    ws[f"L{site_subtotal_row}"].number_format = '#,##0.00'
    ws[f"L{site_subtotal_row}"].font = Font(bold=True)
    
    # GRAND TOTAL
    grand_total_row = site_subtotal_row + 2
    ws[f"B{grand_total_row}"] = "GRAND TOTAL"
    ws[f"B{grand_total_row}"].font = Font(bold=True, size=14)
    ws[f"B{grand_total_row}"].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    ws[f"L{grand_total_row}"] = f"=L{subtotal_row}+L{site_subtotal_row}"
    ws[f"L{grand_total_row}"].number_format = '#,##0.00'
    ws[f"L{grand_total_row}"].font = Font(bold=True, size=14)
    ws[f"L{grand_total_row}"].fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")

    # Apply borders to all data rows
    for row in range(8, grand_total_row + 1):
        for col in ['B', 'G', 'H', 'I', 'J', 'L']:
            ws[f"{col}{row}"].border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

    # Set column widths
    column_widths = {
        'A': 2, 'B': 50, 'C': 30, 'D': 8, 'E': 8, 'F': 8,
        'G': 12, 'H': 12, 'I': 18, 'J': 18, 'K': 8, 'L': 18
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze panes at row 10 (after headers)
    ws.freeze_panes = "B10"

    # Add instructions
    ws["A1"] = "← Enter project name here"
    ws["A2"] = "← Enter location here"
    ws["A3"] = "← Enter work type here"
    ws["A8"] = "← Use standard UOM: sq.m, cu.m, pcs, kg, lot, lm"
    ws["A10"] = "← Section headers help organize work"

    # Save the file
    wb.save("powermason_capstone - Copy/static/templates/Company_BOQ_Template.xlsx")
    print("Company BOQ Template Excel file created successfully!")

if __name__ == "__main__":
    create_company_boq_template()
